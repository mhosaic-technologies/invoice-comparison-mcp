"""
Comparison engine for invoice product analysis
"""

import csv
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict, field
from io import StringIO, BytesIO
from datetime import datetime
from invoice_comparison.matching.product_matcher import ProductMatcher, MatchResult
from invoice_comparison.utils import normalize_gtin, MatchStatus, MatchType, match_status_to_display
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def sanitize_excel_value(value):
    """
    Sanitize value to prevent Excel formula injection

    Excel treats cells starting with =, +, -, @, | as formulas.
    This can be exploited for command execution (CSV/Excel injection).

    Solution: Prefix dangerous values with a single quote to force text interpretation.

    Args:
        value: Cell value (any type)

    Returns:
        Sanitized value safe for Excel
    """
    # Only sanitize string values
    if not isinstance(value, str):
        return value

    # If empty string, return as-is
    if not value:
        return value

    # Check if starts with dangerous characters
    dangerous_chars = ('=', '+', '-', '@', '|')
    if value[0] in dangerous_chars:
        # Prefix with single quote to force Excel to treat as text
        return "'" + value

    return value


@dataclass
class InvoiceItem:
    """Single item from invoice CSV"""
    supplier_code: str
    product_name: str
    brand: str
    format: str
    packaging: str
    category: str
    price: float
    quantity: float

    @property
    def line_total(self) -> float:
        """Calculate line item total"""
        return self.price * self.quantity

    def to_dict(self) -> Dict:
        """Convert to dictionary for matching"""
        return {
            'supplier_code': self.supplier_code,
            'product_name': self.product_name,
            'brand': self.brand,
            'format': self.format,
            'packaging': self.packaging
        }


@dataclass
class ComparisonResult:
    """Result of comparing a single invoice item"""
    # Original invoice item
    original_item: InvoiceItem

    # Source product (if found in database by supplier code)
    source_product: Optional[object] = None  # Product object

    # Match results from target supplier
    matches: List[MatchResult] = field(default_factory=list)

    # Best match (if any)
    best_match: Optional[MatchResult] = None

    # Price comparison (if match found)
    price_difference: Optional[float] = None
    savings_amount: Optional[float] = None
    savings_percent: Optional[float] = None

    # Status
    match_status: str = MatchStatus.NO_MATCH  # Uses MatchStatus constants

    def __post_init__(self):
        """Calculate comparisons after initialization"""
        if self.matches and len(self.matches) > 0:
            self.best_match = self.matches[0]

            # Determine match status
            if self.best_match.match_type == 'gtin':
                self.match_status = MatchStatus.EXACT_MATCH
            elif self.best_match.similarity_score >= 80.0:
                self.match_status = MatchStatus.FUZZY_MATCH
            else:
                # Any other match (passed min_similarity threshold) is at least low confidence
                # This handles cases where min_similarity < 60% (e.g., 50%)
                self.match_status = MatchStatus.LOW_CONFIDENCE

            # Calculate price comparison (only if both prices are available and non-zero)
            if (self.best_match.price is not None and self.best_match.price > 0 and
                self.original_item.price is not None and self.original_item.price > 0):
                target_price = self.best_match.price
                original_price = self.original_item.price

                self.price_difference = original_price - target_price
                self.savings_amount = self.price_difference * self.original_item.quantity
                self.savings_percent = (self.price_difference / original_price) * 100


@dataclass
class ComparisonReport:
    """Complete comparison report for an invoice"""
    source_supplier: str
    target_supplier: str
    results: List[ComparisonResult]

    # Summary statistics
    total_items: int = 0
    matched_items: int = 0
    exact_matches: int = 0
    fuzzy_matches: int = 0
    low_confidence_matches: int = 0
    no_matches: int = 0

    # Financial summary
    original_total: float = 0.0
    target_total: float = 0.0
    potential_savings: float = 0.0
    savings_percent: float = 0.0

    def __post_init__(self):
        """Calculate summary statistics"""
        self.total_items = len(self.results)

        for result in self.results:
            # Count matches by type
            if result.match_status == MatchStatus.EXACT_MATCH:
                self.exact_matches += 1
                self.matched_items += 1
            elif result.match_status == MatchStatus.FUZZY_MATCH:
                self.fuzzy_matches += 1
                self.matched_items += 1
            elif result.match_status == MatchStatus.LOW_CONFIDENCE:
                # Low confidence matches are still matches (passed min_similarity threshold)
                self.low_confidence_matches += 1
                self.matched_items += 1
            elif result.match_status == MatchStatus.NO_MATCH:
                self.no_matches += 1

            # Calculate financial totals
            self.original_total += result.original_item.line_total

            if result.best_match and result.best_match.price:
                self.target_total += result.best_match.price * result.original_item.quantity

                if result.savings_amount:
                    self.potential_savings += result.savings_amount

        # Calculate overall savings percentage
        if self.original_total > 0:
            self.savings_percent = (self.potential_savings / self.original_total) * 100

    def to_dict(self) -> Dict:
        """Convert report to dictionary"""
        return {
            'source_supplier': self.source_supplier,
            'target_supplier': self.target_supplier,
            'summary': {
                'total_items': self.total_items,
                'matched_items': self.matched_items,
                'exact_matches': self.exact_matches,
                'fuzzy_matches': self.fuzzy_matches,
                'low_confidence_matches': self.low_confidence_matches,
                'no_matches': self.no_matches,
                'match_rate': f"{(self.matched_items / self.total_items * 100):.1f}%" if self.total_items > 0 else "0%"
            },
            'financials': {
                'original_total': round(self.original_total, 2),
                'target_total': round(self.target_total, 2),
                'potential_savings': round(self.potential_savings, 2),
                'savings_percent': round(self.savings_percent, 2)
            },
            'items': [
                {
                    'original': {
                        'code': r.original_item.supplier_code,
                        'name': r.original_item.product_name,
                        'brand': r.original_item.brand,
                        'format': r.original_item.format,
                        'price': r.original_item.price,
                        'quantity': r.original_item.quantity,
                        'total': round(r.original_item.line_total, 2),
                        'gtin': r.source_product.gtin if r.source_product else None
                    },
                    'match': {
                        'status': r.match_status,
                        'similarity': round(r.best_match.similarity_score, 1) if r.best_match else 0,
                        'match_type': r.best_match.match_type if r.best_match else None,
                        'product': {
                            'name': r.best_match.product.product_name if r.best_match else None,
                            'brand': r.best_match.product.brand if r.best_match else None,
                            'format': r.best_match.product.format if r.best_match else None,
                            'gtin': r.best_match.product.gtin if r.best_match else None,
                            'code': r.best_match.supplier_code if r.best_match else None,
                            'price': r.best_match.price if r.best_match else None
                        } if r.best_match else None,
                        'price_comparison': {
                            'difference': round(r.price_difference, 2) if r.price_difference else None,
                            'savings': round(r.savings_amount, 2) if r.savings_amount else None,
                            'savings_percent': round(r.savings_percent, 1) if r.savings_percent else None
                        } if r.price_difference else None
                    },
                    'alternatives': [
                        {
                            'similarity': round(m.similarity_score, 1),
                            'name': m.product.product_name,
                            'brand': m.product.brand,
                            'format': m.product.format,
                            'code': m.supplier_code,
                            'price': m.price
                        }
                        for m in r.matches[1:5]  # Include up to 4 alternatives
                    ] if len(r.matches) > 1 else []
                }
                for r in self.results
            ]
        }

    def to_excel_bytes(self) -> bytes:
        """
        Generate Excel file as bytes with GTIN as first column

        Returns:
            Excel file content as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Comparison"

        # Define colors
        exact_match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        fuzzy_match_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
        no_match_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        header_font = Font(bold=True, color="FFFFFF")

        # Column headers - GTIN FIRST
        # Note: Old/New Target Price pattern - Old shows DB price (reference), New is blank for updates
        # Note: Single GTIN column since matched products share the same GTIN
        headers = [
            "GTIN",
            "Source Code",
            "Product Name",
            "Brand",
            "Format",
            "Packaging",
            "Category",
            "Source Price",
            "Quantity",
            "Line Total",
            "Old Target Price",  # From database (read-only reference)
            "New Target Price",  # Empty column for user to fill in updates
            "Match Type",  # Must match what read_comparison_file expects
            "Similarity %",
            "Target Code",
            "Target Product",
            "Target Brand",
            "Target Format"
        ]

        # Identify currency columns by name (more robust than hardcoding indices)
        currency_columns = []
        for idx, header in enumerate(headers, 1):
            if header in ["Source Price", "Line Total", "Old Target Price", "New Target Price", "Price Difference"]:
                currency_columns.append(idx)

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        row_num = 2
        for result in self.results:
            item = result.original_item
            best_match = result.best_match

            # Determine row color based on match status
            if result.match_status == MatchStatus.EXACT_MATCH:
                row_fill = exact_match_fill
            elif result.match_status in [MatchStatus.FUZZY_MATCH, MatchStatus.LOW_CONFIDENCE]:
                row_fill = fuzzy_match_fill
            else:
                row_fill = no_match_fill

            # Build row data - GTIN FIRST (keep source prices, remove comparison columns)
            # Show GTIN from best match if available, otherwise from source product
            gtin_value = ""
            if best_match:
                gtin_value = best_match.product.gtin
            elif result.source_product:
                gtin_value = result.source_product.gtin

            # Sanitize all text values to prevent Excel formula injection
            row_data = [
                sanitize_excel_value(gtin_value),  # GTIN from match or source product
                sanitize_excel_value(item.supplier_code),
                sanitize_excel_value(item.product_name),
                sanitize_excel_value(item.brand),
                sanitize_excel_value(item.format),
                sanitize_excel_value(item.packaging),
                sanitize_excel_value(item.category),
                item.price if item.price else "",  # Source price from invoice (numeric, safe)
                item.quantity if item.quantity else "",  # Numeric, safe
                item.line_total if item.price and item.quantity else "",  # Numeric, safe
                best_match.price if best_match and best_match.price else "",  # Old Target Price (from DB, read-only reference)
                "",  # New Target Price (empty for user to fill in updates)
                sanitize_excel_value(match_status_to_display(result.match_status)),
                f"{best_match.similarity_score:.1f}" if best_match else "0.0",  # Numeric string, safe
                sanitize_excel_value(best_match.supplier_code if best_match else ""),
                sanitize_excel_value(best_match.product.product_name if best_match else ""),
                sanitize_excel_value(best_match.product.brand if best_match else ""),
                sanitize_excel_value(best_match.product.format if best_match else "")
            ]

            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Format currency columns
                if col_num in currency_columns:
                    if value is not None and value != "":
                        cell.number_format = '$#,##0.00'

            row_num += 1

        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])

            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (AttributeError, TypeError):
                    # Skip cells with values that can't be converted to string
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary at the top (before data)
        ws.insert_rows(1, 3)

        # Summary information
        ws['A1'] = f"Invoice Comparison: {self.source_supplier.replace('_', ' ').title()} → {self.target_supplier.replace('_', ' ').title()}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        match_rate = (self.matched_items / self.total_items * 100) if self.total_items > 0 else 0
        ws['A3'] = f"Match Rate: {match_rate:.1f}% ({self.matched_items}/{self.total_items} products)"

        # Freeze summary rows + header row (rows 1-4)
        ws.freeze_panes = "A5"

        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()


class ComparisonEngine:
    """Main comparison engine"""

    def __init__(self, db_path: str = None):
        """Initialize comparison engine"""
        self.matcher = ProductMatcher(db_path)

    def parse_csv(self, csv_content: str) -> List[InvoiceItem]:
        """
        Parse CSV content into invoice items

        Expected CSV format:
        supplier_code,product_name,brand,format,packaging,category,price,quantity

        Args:
            csv_content: CSV string content

        Returns:
            List of InvoiceItem objects
        """
        items = []

        # Parse CSV
        csv_file = StringIO(csv_content)
        reader = csv.DictReader(csv_file)

        # Enumerate rows starting from 2 (row 1 is headers)
        for row_num, row in enumerate(reader, start=2):
            try:
                # Get raw values for error reporting
                price_raw = row.get('price', '').strip()
                quantity_raw = row.get('quantity', '').strip()

                # Parse price with detailed error handling
                try:
                    if not price_raw:
                        # Empty price - default to 0
                        price = 0.0
                    else:
                        # Remove common currency symbols and formatting
                        price_cleaned = price_raw.replace('$', '').replace(',', '').strip()
                        price = float(price_cleaned)
                except ValueError:
                    print(f"Warning: Row {row_num}: Invalid price '{price_raw}' - must be a number (e.g., '10.50' not '$10.50'). Skipping row.")
                    continue

                # Parse quantity with detailed error handling
                try:
                    if not quantity_raw:
                        # Empty quantity - default to 0
                        quantity = 0.0
                    else:
                        quantity = float(quantity_raw)
                except ValueError:
                    print(f"Warning: Row {row_num}: Invalid quantity '{quantity_raw}' - must be a number (e.g., '5' or '2.5'). Skipping row.")
                    continue

                # Validate non-negative values
                if price < 0:
                    print(f"Warning: Row {row_num}: Negative price ({price}) not allowed. Skipping row.")
                    continue
                if quantity < 0:
                    print(f"Warning: Row {row_num}: Negative quantity ({quantity}) not allowed. Skipping row.")
                    continue

                # Check for missing required fields
                supplier_code = row.get('supplier_code', '').strip()
                product_name = row.get('product_name', '').strip()

                if not supplier_code and not product_name:
                    print(f"Warning: Row {row_num}: Missing both supplier code and product name. Skipping row.")
                    continue

                item = InvoiceItem(
                    supplier_code=supplier_code,
                    product_name=product_name,
                    brand=row.get('brand', '').strip(),
                    format=row.get('format', '').strip(),
                    packaging=row.get('packaging', '').strip(),
                    category=row.get('category', '').strip(),
                    price=price,
                    quantity=quantity
                )
                items.append(item)
            except KeyError as e:
                # Missing required column
                print(f"Warning: Row {row_num}: Missing required column {e}. Check CSV headers match expected format.")
                continue
            except Exception as e:
                # Catch-all for unexpected errors
                print(f"Warning: Row {row_num}: Unexpected error - {type(e).__name__}: {e}. Skipping row.")
                continue

        return items

    def compare_invoice(
        self,
        csv_content: str,
        source_supplier: str,
        target_supplier: str,
        min_similarity: float = 60.0,
        max_alternatives: int = 5
    ) -> ComparisonReport:
        """
        Compare invoice products against target supplier

        Args:
            csv_content: CSV string with invoice data
            source_supplier: Source supplier code (e.g., 'dube_loiselle')
            target_supplier: Target supplier code (e.g., 'colabor')
            min_similarity: Minimum similarity threshold for fuzzy matching
            max_alternatives: Maximum alternative matches to return

        Returns:
            ComparisonReport with detailed results
        """
        # Parse CSV
        items = self.parse_csv(csv_content)

        if not items:
            raise ValueError("No valid items found in CSV")

        # Process each item
        results = []

        for item in items:
            # Try to look up source product by supplier code (to get GTIN)
            source_product = None
            if item.supplier_code:
                source_result = self.matcher.db_ops.find_product_by_supplier_code(
                    item.supplier_code,
                    source_supplier
                )
                if source_result:
                    source_product, _ = source_result

            # Find matches
            matches = self.matcher.find_matches(
                product_info=item.to_dict(),
                source_supplier=source_supplier,
                target_supplier=target_supplier,
                min_similarity=min_similarity,
                max_results=max_alternatives
            )

            # Create comparison result
            result = ComparisonResult(
                original_item=item,
                source_product=source_product,
                matches=matches
            )

            results.append(result)

        # Create report
        report = ComparisonReport(
            source_supplier=source_supplier,
            target_supplier=target_supplier,
            results=results
        )

        return report

    def compare_from_file(
        self,
        csv_file_path: str,
        source_supplier: str,
        target_supplier: str,
        min_similarity: float = 60.0
    ) -> ComparisonReport:
        """
        Compare invoice from CSV file

        Args:
            csv_file_path: Path to CSV file
            source_supplier: Source supplier code
            target_supplier: Target supplier code
            min_similarity: Minimum similarity threshold

        Returns:
            ComparisonReport
        """
        with open(csv_file_path, 'r', encoding='utf-8') as f:
            csv_content = f.read()

        return self.compare_invoice(
            csv_content,
            source_supplier,
            target_supplier,
            min_similarity
        )

    def import_corrections(
        self,
        csv_content: str,
        source_supplier: str,
        target_supplier: str
    ) -> Dict:
        """
        Import corrections from edited Excel/CSV file

        This function now CREATES products and supplier codes if they don't exist.
        Any new information is valuable knowledge and will be added to the database.

        Expected CSV columns:
        - GTIN (optional but recommended)
        - Source Code (required)
        - Product Name (optional but recommended)
        - Brand (optional)
        - Format (optional)
        - Target Code (optional - if provided, creates mapping)

        Args:
            csv_content: CSV string with corrections
            source_supplier: Source supplier code (e.g., 'dube_loiselle')
            target_supplier: Target supplier code (e.g., 'colabor')

        Returns:
            Dict with summary of products/corrections saved
        """
        from invoice_comparison.database.schema import Product, SupplierCode, Supplier
        from datetime import datetime

        db = self.matcher.db_ops

        products_created = []
        products_updated = []
        corrections_saved = []
        corrections_failed = []
        corrections_skipped = []

        # Parse CSV
        csv_file = StringIO(csv_content)
        reader = csv.DictReader(csv_file)

        session = db.get_session()

        try:
            # Get supplier objects
            source_supplier_obj = session.query(Supplier).filter_by(code=source_supplier).first()
            target_supplier_obj = session.query(Supplier).filter_by(code=target_supplier).first()

            if not source_supplier_obj or not target_supplier_obj:
                return {
                    'products_created': [],
                    'products_updated': [],
                    'saved': [],
                    'failed': [{'reason': 'Supplier not found in database'}],
                    'skipped': [],
                    'summary': {
                        'total_products_created': 0,
                        'total_products_updated': 0,
                        'total_saved': 0,
                        'total_failed': 1,
                        'total_skipped': 0
                    }
                }

            for row in reader:
                source_code = row.get('Source Code', '').strip()
                target_code = row.get('Target Code', '').strip()
                gtin_raw = row.get('GTIN', '').strip()
                product_name = row.get('Product Name', '').strip()
                brand = row.get('Brand', '').strip() or None
                format_val = row.get('Format', '').strip() or None
                packaging = row.get('Packaging', '').strip() or None
                category = row.get('Category', '').strip() or None

                # Read price information (prices change frequently, so we update them)
                # Note: We read target price from multiple possible column names for flexibility
                # Try multiple column name variations (case-insensitive)
                def get_price_from_row(row, *column_variations):
                    """Try multiple column name variations to find price value"""
                    for col_name in column_variations:
                        # Try exact match first
                        if col_name in row:
                            return row.get(col_name, '').strip()
                        # Try case-insensitive match
                        for key in row.keys():
                            if key.lower() == col_name.lower():
                                return row.get(key, '').strip()
                    return ''

                # Source price variations
                source_price_str = get_price_from_row(row, 'Source Price', 'Price Source')

                # Target price variations (most to least specific)
                target_price_str = get_price_from_row(
                    row,
                    'New Target Price',      # Our standard column name
                    'Target Price',          # User might use this
                    'Price',                 # Generic price column
                    f'{target_supplier.title()} Price',  # Supplier-specific (e.g., "Colabor Price")
                    target_supplier.replace('_', ' ').title() + ' Price'  # With spaces
                )

                # Parse prices (handle empty strings and convert to float)
                source_price = None
                target_price = None
                try:
                    if source_price_str:
                        source_price = float(source_price_str)
                except (ValueError, TypeError):
                    pass  # Invalid price format, ignore

                try:
                    if target_price_str:
                        target_price = float(target_price_str)
                except (ValueError, TypeError):
                    pass  # Invalid price format, ignore

                # Skip if no source code
                if not source_code:
                    continue

                # Normalize and validate GTIN (handles "12345.0" -> "12345" and validates format)
                gtin = normalize_gtin(gtin_raw) if gtin_raw else None

                try:
                    # Step 1: GTIN is REQUIRED for creating new mappings
                    if not gtin:
                        corrections_failed.append({
                            'source_code': source_code,
                            'product_name': product_name or 'N/A',
                            'reason': 'GTIN is required. Please provide a GTIN in column A of the Excel file.'
                        })
                        continue

                    # Step 2: Find or create product using the EXACT GTIN provided
                    product = session.query(Product).filter_by(gtin=gtin).first()
                    product_created = False
                    product_updated = False

                    if not product:
                        # Create new product with user's GTIN
                        product = Product(
                            gtin=gtin,
                            product_name=product_name or f"Product {source_code}",
                            brand=brand,
                            format=format_val,
                            packaging=packaging,
                            category=category,
                            created_at=datetime.utcnow()
                        )
                        session.add(product)
                        session.flush()  # Get product ID
                        product_created = True
                        products_created.append({
                            'gtin': gtin,
                            'product_name': product.product_name,
                            'source_code': source_code
                        })
                    else:
                        # Update existing product when user provides different values (corrections)
                        updated_fields = []
                        if product_name and product.product_name != product_name:
                            product.product_name = product_name
                            updated_fields.append('product_name')
                        if brand and product.brand != brand:
                            product.brand = brand
                            updated_fields.append('brand')
                        if format_val and product.format != format_val:
                            product.format = format_val
                            updated_fields.append('format')
                        if packaging and product.packaging != packaging:
                            product.packaging = packaging
                            updated_fields.append('packaging')
                        if category and product.category != category:
                            product.category = category
                            updated_fields.append('category')

                        if updated_fields:
                            product.updated_at = datetime.utcnow()
                            product_updated = True
                            products_updated.append({
                                'gtin': product.gtin,
                                'product_name': product.product_name,
                                'updated_fields': updated_fields
                            })

                    # Step 3: Create or update source supplier code mapping
                    source_supplier_code = session.query(SupplierCode).filter_by(
                        supplier_id=source_supplier_obj.id,
                        supplier_code=source_code
                    ).first()

                    if source_supplier_code:
                        # Supplier code exists - check if it points to the same product
                        if source_supplier_code.product_id != product.id:
                            # Conflict: Same supplier code mapped to different product
                            # This could happen if GTIN changed or data was corrected
                            # Update to point to the new product (trust user's GTIN)
                            source_supplier_code.product_id = product.id
                            source_supplier_code.active = True
                            source_supplier_code.updated_at = datetime.utcnow()

                        # Update price if provided (prices change frequently)
                        if source_price is not None:
                            source_supplier_code.price = source_price
                            source_supplier_code.price_updated_at = datetime.utcnow()
                    else:
                        # Create new supplier code mapping
                        source_supplier_code = SupplierCode(
                            supplier_id=source_supplier_obj.id,
                            product_id=product.id,
                            supplier_code=source_code,
                            price=source_price,  # Set initial price if provided
                            price_updated_at=datetime.utcnow() if source_price is not None else None,
                            active=True,
                            created_at=datetime.utcnow()
                        )
                        session.add(source_supplier_code)

                    # Step 4: If target code provided, find or create target product mapping
                    if target_code:
                        # Try to find target product by code (pass session to avoid nested sessions)
                        target_result = db.find_product_by_supplier_code(target_code, target_supplier, session=session)

                        if target_result:
                            # Target product exists - verify GTIN consistency
                            target_product, target_supplier_code_obj = target_result

                            # Check if GTINs match
                            if target_product.gtin != product.gtin:
                                # GTIN conflict: user provided one GTIN, but target code is linked to different GTIN
                                # This indicates a data inconsistency that the user should resolve
                                corrections_failed.append({
                                    'source_code': source_code,
                                    'target_code': target_code,
                                    'product_name': product_name,
                                    'reason': f'GTIN conflict: User provided GTIN "{product.gtin}" but target code "{target_code}" is linked to product with GTIN "{target_product.gtin}". Please verify the correct GTIN and update the Excel file.'
                                })
                                # Skip this row but continue processing others (don't rollback)
                                continue

                            # GTINs match - update target price if provided (prices change frequently)
                            if target_price is not None:
                                target_supplier_code_obj.price = target_price
                                target_supplier_code_obj.price_updated_at = datetime.utcnow()

                            # Create correction/mapping
                            correction_data = {
                                'original_supplier_id': source_supplier_obj.id,
                                'original_supplier_code': source_code,
                                'original_description': product_name or product.product_name,
                                'original_format': format_val or product.format,
                                'matched_product_id': product.id,
                                'target_supplier_id': target_supplier_obj.id,
                                'target_supplier_code': target_code,
                                'similarity_score': 100.0,
                                'user_confirmed': True
                            }

                            # Pass session to avoid nested sessions and commits
                            db.add_user_correction(correction_data, session=session)

                            corrections_saved.append({
                                'source_code': source_code,
                                'source_product': product.product_name,
                                'target_code': target_code,
                                'target_product': product.product_name,
                                'gtin': product.gtin,
                                'product_created': product_created,
                                'product_updated': product_updated
                            })
                        else:
                            # Target product doesn't exist - create or update supplier code mapping
                            # Check if this target code already exists (safety check)
                            existing_target_code = session.query(SupplierCode).filter_by(
                                supplier_id=target_supplier_obj.id,
                                supplier_code=target_code
                            ).first()

                            if existing_target_code:
                                # Update existing mapping to point to new product
                                existing_target_code.product_id = product.id
                                existing_target_code.active = True
                                existing_target_code.updated_at = datetime.utcnow()

                                # Update target price if provided (prices change frequently)
                                if target_price is not None:
                                    existing_target_code.price = target_price
                                    existing_target_code.price_updated_at = datetime.utcnow()
                            else:
                                # Create new supplier code mapping
                                target_supplier_code = SupplierCode(
                                    supplier_id=target_supplier_obj.id,
                                    product_id=product.id,
                                    supplier_code=target_code,
                                    price=target_price,  # Set initial price if provided
                                    price_updated_at=datetime.utcnow() if target_price is not None else None,
                                    active=True,
                                    created_at=datetime.utcnow()
                                )
                                session.add(target_supplier_code)

                            corrections_saved.append({
                                'source_code': source_code,
                                'source_product': product.product_name,
                                'target_code': target_code,
                                'target_product': product.product_name,
                                'gtin': product.gtin,
                                'product_created': product_created,
                                'product_updated': product_updated,
                                'note': 'Target code added to same product'
                            })

                    # Commit after each successful row to prevent cascading failures
                    session.commit()

                except Exception as e:
                    # Rollback only this row's changes, not the entire import
                    session.rollback()
                    corrections_failed.append({
                        'source_code': source_code,
                        'target_code': target_code,
                        'product_name': product_name,
                        'reason': str(e)
                    })

        finally:
            session.close()

        return {
            'products_created': products_created,
            'products_updated': products_updated,
            'saved': corrections_saved,
            'failed': corrections_failed,
            'skipped': corrections_skipped,
            'summary': {
                'total_products_created': len(products_created),
                'total_products_updated': len(products_updated),
                'total_saved': len(corrections_saved),
                'total_failed': len(corrections_failed),
                'total_skipped': len(corrections_skipped)
            }
        }


if __name__ == "__main__":
    # Test the comparison engine
    print("=== Testing Comparison Engine ===\n")

    engine = ComparisonEngine()

    # Sample CSV data (from invoice)
    sample_csv = """supplier_code,product_name,brand,format,packaging,category,price,quantity
325141,YOGOURT VANILLE 1.5% ORIG IOGO,IOGO,4X2 KG,,600,43.90,1
162609,TOFU FERME BIO SOUS VIDE,,12X454 G,,600,35.15,2
238533,CEREALE CHEERIOS VRAC,CHEERIOS,4X822 G,,100,42.30,1
155915,OEUF REF VRAC LARGE,,15 DOUZ,,600,59.05,1"""

    print("Sample CSV:")
    print(sample_csv)
    print("\n" + "=" * 80 + "\n")

    # Run comparison
    report = engine.compare_invoice(
        csv_content=sample_csv,
        source_supplier="dube_loiselle",
        target_supplier="colabor",
        min_similarity=60.0
    )

    # Print summary
    print(f"Comparison: {report.source_supplier} → {report.target_supplier}\n")
    print(f"Total items: {report.total_items}")
    if report.total_items > 0:
        print(f"Matched: {report.matched_items} ({report.matched_items/report.total_items*100:.0f}%)")
    else:
        print(f"Matched: {report.matched_items} (0%)")
    print(f"  - Exact matches: {report.exact_matches}")
    print(f"  - Fuzzy matches: {report.fuzzy_matches}")
    print(f"  - Low confidence matches: {report.low_confidence_matches}")
    print(f"  - No matches: {report.no_matches}\n")

    print(f"Original total: ${report.original_total:.2f}")
    print(f"Target total: ${report.target_total:.2f}")
    print(f"Potential savings: ${report.potential_savings:.2f} ({report.savings_percent:.1f}%)\n")

    print("=" * 80)
    print("DETAILED RESULTS")
    print("=" * 80 + "\n")

    for i, result in enumerate(report.results, 1):
        print(f"{i}. {result.original_item.product_name}")
        print(f"   Original: ${result.original_item.price:.2f} × {result.original_item.quantity:.0f} = ${result.original_item.line_total:.2f}")

        if result.best_match:
            print(f"   Match: {result.best_match.product.product_name}")
            print(f"   Similarity: {result.best_match.similarity_score:.1f}% ({result.match_status})")

            if result.best_match.price:
                print(f"   Target: ${result.best_match.price:.2f} × {result.original_item.quantity:.0f} = ${result.best_match.price * result.original_item.quantity:.2f}")

                if result.savings_amount:
                    sign = "+" if result.savings_amount > 0 else ""
                    print(f"   Savings: {sign}${result.savings_amount:.2f} ({sign}{result.savings_percent:.1f}%)")
            else:
                print(f"   Target: Price not available")
        else:
            print(f"   Match: Not found at {report.target_supplier}")

        print()

    print("=== Test Complete ===")
